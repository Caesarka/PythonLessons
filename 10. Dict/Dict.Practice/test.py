from openpyxl import Workbook, load_workbook

wb = load_workbook('grades.xlsx')
sheet = wb.active
sheet.title = "Grades"
new_sheet = wb.create_sheet('Summary')

avg_score_student = {}
avg_score_course = {}
score_course = {}

max_score_stud = 0
max_score_students = []
min_score_courses = []


rows = sheet.iter_rows(min_row=1, values_only=True)
grades = [list(row) for row in rows]
course_cnt = len(grades[0]) - 1
student_cnt = len(grades) - 1

for row in range(1, len(grades)):
    avg_score = 0

    for col in range(1, len(grades[row])):
        avg_score += grades[row][col]
        score_course.setdefault(grades[0][col], 0)
        score_course[grades[0][col]] += grades[row][col]

    avg_score_student[grades[row][0]] = round(avg_score / course_cnt, 2)

    if max_score_stud < avg_score_student[grades[row][0]]:
        max_score_stud = avg_score_student[grades[row][0]]

avg_score_course = {key: value / student_cnt for key,
                    value in score_course.items()}

max_score_students = ', '.join([
    key for key, value in avg_score_student.items() if value == max_score_stud])

min_score_cour = min(avg_score_course.values())

min_score_courses = ', '.join([
    key for key, value in avg_score_course.items() if value == min_score_cour])

new_sheet.cell(row=1, column=1).value = 'Student name'
new_sheet.cell(row=1, column=2).value = 'Average score'
r = 2
c = 1
for key, value in avg_score_student.items():
    new_sheet.cell(row=r, column=c).value = key
    new_sheet.cell(row=r, column=c + 1).value = value
    r += 1

new_sheet.cell(row=1, column=4).value = 'Course name'
new_sheet.cell(row=1, column=5).value = 'Average score'
r = 2
c = 4
for key, value in avg_score_course.items():
    new_sheet.cell(row=r, column=c).value = key
    new_sheet.cell(row=r, column=c + 1).value = value
    r += 1

new_sheet.cell(row=1, column=7).value = 'Best student score'
new_sheet.cell(row=2, column=7).value = max_score_stud
new_sheet.cell(row=1, column=8).value = 'Students'
new_sheet.cell(row=2, column=8).value = max_score_students

new_sheet.cell(row=1, column=10).value = 'Lowest course score'
new_sheet.cell(row=2, column=10).value = min_score_cour
new_sheet.cell(row=1, column=11).value = 'Courses'
new_sheet.cell(row=2, column=11).value = min_score_courses


wb.save('grades.xlsx')
